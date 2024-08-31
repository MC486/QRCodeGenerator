import requests
import segno
from PIL import Image, ImageChops
import openpyxl
from openpyxl.drawing.image import Image as ExcelImage


def read_auth_token(file_path):
    with open(file_path, 'r') as file:
        token = file.read().strip()  # Read the token and remove any surrounding whitespace
        token = "Bearer " + token  # Add the 'Bearer ' prefix
    print(f"Token read from file: '{token}'")  # Print the token to the console for debugging

    return token


def shorten_url_with_tinyurl(url):
    auth_token = read_auth_token('tinyurl_auth_token.txt')
    api_url = "https://api.tinyurl.com/create"
    headers = {
        "Content-Type": "application/json",
        "Authorization": auth_token
    }
    data = {
        "url": url,
        "domain": "tinyurl.com"
    }
    response = requests.post(api_url, json=data, headers=headers)

    if response.status_code == 200:
        return response.json().get("data", {}).get("tiny_url")
    else:
        print(f"Error: Could not shorten URL. Status Code: {response.status_code}")
        print(response.json())
        return url  # Return original URL if shortening fails


def save_to_excel(file_name, xlsx_user_input, xlsx_original_url, xlsx_short_url, xlsx_qr_img):
    # Check if the file exists, if not create it with headers
    try:
        workbook = openpyxl.load_workbook(file_name)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet['A1'] = "User Input"
        sheet['B1'] = "TinyURL"
        sheet['C1'] = "Original URL"
        workbook.save(file_name)

    # Load the workbook and get the active sheet
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook.active

    # Find the next empty row in the Excel sheet
    next_row = sheet.max_row + 1

    # Write the data into the Excel sheet
    sheet[f'A{next_row}'] = xlsx_user_input
    sheet[f'B{next_row}'] = xlsx_short_url
    sheet[f'C{next_row}'] = xlsx_original_url

    # Add the QR code image to the sheet
    img = ExcelImage(xlsx_qr_img)
    img.width = 64
    img.height = 64
    img_anchor = f'D{next_row}'  # Anchor the image in the D column at the next available row
    sheet.add_image(img, img_anchor)

    # Save the workbook
    workbook.save(file_name)


def generate_qr_code_with_high_error_correction(data, scale=10, include_logo=False, logo_path=None):
    qr = segno.make(data, error='h')  # High error correction
    qr_file = "temp_qr_with_logo.png" if include_logo else "temp_qr_without_logo.png"
    qr.save(qr_file, scale=scale)

    if include_logo and logo_path:
        qr_img = Image.open(qr_file).convert("RGBA")
        logo = Image.open(logo_path).convert("RGBA")
        logo_size = qr_img.size[0] // 10  # Resize logo to 10% of QR code size
        logo = logo.resize((logo_size, logo_size), Image.Resampling.LANCZOS)
        pos_center = ((qr_img.size[0] - logo_size) // 2, (qr_img.size[1] - logo_size) // 2)
        qr_img.paste(logo, pos_center, mask=logo)
        qr_img.save(qr_file)

    return qr_file


def crop_whitespace(image_path):
    img = Image.open(image_path)
    bg = Image.new(img.mode, img.size, img.getpixel((0, 0)))
    diff = ImageChops.difference(img, bg)
    bbox = diff.getbbox()
    if bbox:
        img = img.crop(bbox)
    return img


def tile_qr_code(cropped_img, tile_count=3):
    width, height = cropped_img.size
    tiled_img = Image.new('RGBA', (width * tile_count, height * tile_count), (255, 255, 255, 0))

    for i in range(tile_count):
        for j in range(tile_count):
            tiled_img.paste(cropped_img, (i * width, j * height))

    return tiled_img


def overlay_center_qr(tiled_img, center_qr):
    center_pos = ((tiled_img.size[0] - center_qr.size[0]) // 2, (tiled_img.size[1] - center_qr.size[1]) // 2)
    tiled_img.paste(center_qr, center_pos, mask=center_qr)
    return tiled_img


def main():
    data = input("Enter the URL or data to encode in the QR code: ")
    file_name = input("Enter the file name for the final QR code image (without extension): ")

    # Automatically shorten the URL using TinyURL
    shortened_url = shorten_url_with_tinyurl(data)
    print(f"Shortened URL: {shortened_url}")

    # Paths to images
    template_path = "mickey_template.png"
    logo_path = "sweeping-mickey.png"

    # Step 1: Generate and crop QR code without logo
    qr_without_logo_path = generate_qr_code_with_high_error_correction(shortened_url, scale=30, include_logo=False)
    cropped_qr_without_logo = crop_whitespace(qr_without_logo_path)
    cropped_qr_without_logo.save("cropped_qr_without_logo.png")

    # Step 2: Tile the cropped QR code
    tiled_qr_img = tile_qr_code(cropped_qr_without_logo, tile_count=3)

    # Step 3: Generate QR code with logo
    qr_with_logo_path = generate_qr_code_with_high_error_correction(shortened_url, scale=30, include_logo=True,
                                                                    logo_path=logo_path)

    # Step 4: Crop the whitespace from the QR code with logo
    cropped_qr_with_logo = crop_whitespace(qr_with_logo_path)
    cropped_qr_with_logo.save("cropped_qr_with_logo.png")

    # Step 5: Overlay the cropped center QR code with logo onto the tiled QR code
    final_img = overlay_center_qr(tiled_qr_img, cropped_qr_without_logo)

    # Step 6: Overlay the Mickey Mouse template
    template_img = Image.open(template_path).convert("RGBA")
    final_img = final_img.resize(template_img.size, Image.Resampling.LANCZOS)
    combined_img = Image.alpha_composite(final_img, template_img)

    # Step 7: Save the final image
    output_path = f"{file_name}.png"
    combined_img.save(output_path)
    print(f"Final QR code image saved as {output_path}")

    # Step 8: Save the data to an Excel file
    save_to_excel("tinyurl_links.xlsx", file_name, data, shortened_url, output_path)


if __name__ == "__main__":
    main()
